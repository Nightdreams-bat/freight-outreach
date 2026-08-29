import shutil
import zipfile

import openpyxl
import pytest

from kairo.excel_store import (
    DATA_COLUMNS,
    STATE_COLUMNS,
    ExcelFileLocked,
    ExcelStore,
    sheet_headers,
)


def _make(path, headers, *rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append(list(r))
    wb.save(path)
    return path


def test_auto_detects_aliased_headers(tmp_path):
    p = _make(tmp_path / "leads.xlsx",
              ["Full Name", "Organisation", "E-mail Address", "Mobile"],
              ["Jane Roe", "Acme", "jane@acme.test", "555-1"])
    store = ExcelStore(p)
    rows = list(store.rows())
    assert len(rows) == 1
    _, v = rows[0]
    assert v["Name"] == "Jane Roe"
    assert v["Company"] == "Acme"
    assert v["Email"] == "jane@acme.test"
    assert v["Phone"] == "555-1"


def test_combined_first_last_name(tmp_path):
    p = _make(tmp_path / "leads.xlsx",
              ["First Name", "Last Name", "Email"],
              ["Jane", "Roe", "jane@acme.test"])
    store = ExcelStore(p)
    _, v = list(store.rows())[0]
    assert v["Name"] == "Jane Roe"


def test_save_is_atomic_and_keeps_a_backup(tmp_path):
    p = _make(tmp_path / "leads.xlsx",
              ["Name", "Email"], ["Jane", "jane@acme.test"])
    original = p.read_bytes()

    store = ExcelStore(p)
    # .bak is written once on construction, before any change.
    bak = p.with_suffix(".xlsx.bak")
    assert bak.exists()
    assert bak.read_bytes() == original

    row_idx = list(store.rows())[0][0]
    store.set_value(row_idx, "Notes", "touched")
    # No stray temp file left behind, and the sheet is still readable.
    assert not (tmp_path / "leads.xlsx.tmp").exists()
    reloaded = ExcelStore(p)
    assert reloaded.get_row(row_idx)["Notes"] == "touched"


def test_explicit_map_overrides_detection(tmp_path):
    p = _make(tmp_path / "leads.xlsx",
              ["Contact", "Alt Contact", "Email"],
              ["Wrong", "Right", "x@acme.test"])
    store = ExcelStore(p, column_map={"Name": "Alt Contact"})
    _, v = list(store.rows())[0]
    assert v["Name"] == "Right"


def test_missing_email_column(tmp_path):
    p = _make(tmp_path / "leads.xlsx", ["Name", "Company"], ["Jane", "Acme"])
    store = ExcelStore(p)
    assert store.email_column_missing is True
    assert list(store.rows()) == []
    assert list(store.all_rows()) == []


def test_email_only_sheet_still_works(tmp_path):
    p = _make(tmp_path / "leads.xlsx", ["Email"], ["jane@acme.test"])
    store = ExcelStore(p)
    _, v = list(store.rows())[0]
    assert v["Email"] == "jane@acme.test"
    assert "Name" not in v  # not mapped -> not read


def test_client_data_columns_never_appended(tmp_path):
    p = _make(tmp_path / "leads.xlsx", ["Full Name", "E-mail Address"],
              ["Jane", "jane@acme.test"])
    store = ExcelStore(p)
    store.set_value(2, "Notes", "x")  # first real write persists the state headers
    headers = {c.value for c in openpyxl.load_workbook(p).active[1]}
    # No Phone/Company column was invented for the client...
    assert "Phone" not in headers
    assert "Company" not in headers
    # ...but our own state columns were appended.
    for col in STATE_COLUMNS:
        assert col in headers


def test_construction_does_not_write_missing_state_columns(tmp_path):
    p = _make(tmp_path / "leads.xlsx", ["Name", "Email"], ["Jane", "jane@acme.test"])
    before_bytes = p.read_bytes()
    before_mtime = p.stat().st_mtime
    ExcelStore(p)
    assert p.read_bytes() == before_bytes
    assert p.stat().st_mtime == before_mtime


def test_state_columns_readable_before_any_write(tmp_path):
    p = _make(tmp_path / "leads.xlsx", ["Name", "Email"], ["Jane", "jane@acme.test"])
    store = ExcelStore(p)
    _, v = list(store.rows())[0]
    assert "Suppressed" in v
    assert v["Suppressed"] in (None, "")
    on_disk = [c.value for c in openpyxl.load_workbook(p).active[1]]
    assert "Suppressed" not in on_disk  # still not persisted


def test_reads_survive_a_blank_column_in_the_header_row(tmp_path):
    # A gap between B and D used to shift every later column, so a Note typed
    # into D was read from the empty C and shown as blank on the Leads page.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"], ws["B1"], ws["D1"] = "Name", "Email", "Notes"
    ws["A2"], ws["B2"], ws["D2"] = "Jane", "jane@acme.test", "called Tuesday"
    p = tmp_path / "gappy.xlsx"
    wb.save(p)

    store = ExcelStore(p)
    _, v, _ = list(store.all_rows())[0]
    assert v["Email"] == "jane@acme.test"
    assert v["Notes"] == "called Tuesday"


def test_first_write_persists_state_headers(tmp_path):
    p = _make(tmp_path / "leads.xlsx", ["Name", "Email"], ["Jane", "jane@acme.test"])
    store = ExcelStore(p)
    store.set_value(2, "Suppressed", "yes")
    on_disk = [c.value for c in openpyxl.load_workbook(p).active[1]]
    for col in STATE_COLUMNS:
        assert col in on_disk
    reloaded = ExcelStore(p)
    assert reloaded.get_row(2)["Suppressed"] == "yes"


def test_picks_worksheet_with_email_column(tmp_path):
    wb = openpyxl.Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append(["Region", "Total"])
    summary.append(["North", 5])
    contacts = wb.create_sheet("Contacts")
    contacts.append(["Name", "Email"])
    contacts.append(["Jane", "jane@acme.test"])
    p = tmp_path / "leads.xlsx"
    wb.save(p)

    store = ExcelStore(p)
    assert store.email_column_missing is False
    _, v = list(store.rows())[0]
    assert v["Email"] == "jane@acme.test"


def test_save_permission_error_raises_locked(tmp_path, monkeypatch):
    p = _make(tmp_path / "leads.xlsx", ["Name", "Email"], ["Jane", "jane@acme.test"])
    store = ExcelStore(p)

    def boom(*a, **k):
        raise PermissionError("open in Excel")

    monkeypatch.setattr(store.wb, "save", boom)
    with pytest.raises(ExcelFileLocked):
        store.set_value(2, "Notes", "x")


def test_set_value_refuses_data_column(tmp_path):
    p = _make(tmp_path / "leads.xlsx", ["Name", "Email"], ["Jane", "jane@acme.test"])
    store = ExcelStore(p)
    store.set_value(2, "ReplyStatus", "yes")  # state column: fine
    with pytest.raises(ValueError):
        store.set_value(2, "Name", "Hacked")


def test_sheet_headers_helper(tmp_path):
    p = _make(tmp_path / "leads.xlsx", ["Full Name", "Email", ""], ["a", "b", ""])
    assert sheet_headers(p) == ["Full Name", "Email"]
    assert sheet_headers(tmp_path / "nope.xlsx") == []


def test_add_lead_appends_and_returns_index(tmp_path):
    p = _make(tmp_path / "leads.xlsx", ["Name", "Company", "Email", "Phone"],
              ["Jane", "Acme", "jane@acme.test", "1"])
    store = ExcelStore(p)
    idx = store.add_lead({"Name": "New", "Company": "NewCo", "Email": "new@newco.test"})
    assert idx == 3
    reopened = ExcelStore(p)
    emails = {v["Email"] for _, v, _ in reopened.all_rows()}
    assert "new@newco.test" in emails


def test_add_lead_rejects_duplicate_and_invalid(tmp_path):
    p = _make(tmp_path / "leads.xlsx", ["Name", "Company", "Email", "Phone"],
              ["Jane", "Acme", "jane@acme.test", "1"])
    store = ExcelStore(p)
    assert store.add_lead({"Company": "X", "Email": "JANE@acme.test"}) is None  # dup, case-insensitive
    assert store.add_lead({"Company": "X", "Email": "not-an-email"}) is None
    assert store.add_lead({"Company": "X", "Email": ""}) is None


def test_add_lead_seeds_columns_on_empty_workbook(tmp_path):
    # An operator can make a blank .xlsx by hand in Excel and link it. It has no
    # Name/Email column, so add_lead used to have nowhere to write and every
    # append was silently rejected ("skipped - duplicates / no email").
    p = tmp_path / "blank.xlsx"
    openpyxl.Workbook().save(p)
    store = ExcelStore(p, create_if_missing=False)
    assert not store.email_column_missing
    idx = store.add_lead({"Name": "New", "Company": "NewCo", "Email": "new@newco.test"})
    assert idx is not None
    reopened = ExcelStore(p, create_if_missing=False)
    assert "new@newco.test" in reopened.existing_emails()


def test_state_write_preserves_a_note_typed_in_excel(tmp_path):
    # The operator opens the sheet in Excel and adds a Note while a Kairo store
    # is still alive from earlier. A later state write must merge, not clobber.
    p = _make(tmp_path / "leads.xlsx", ["Name", "Company", "Email", "Phone"],
              ["Jane", "Acme", "jane@acme.test", "1"])
    ExcelStore(p).initialize()  # persist the STATE columns, as a real sheet has
    store = ExcelStore(p)  # loaded now, before the edit

    ext = openpyxl.load_workbook(p)
    ws = ext.active
    notes_col = next(c.column for c in ws[1] if c.value == "Notes")
    ws.cell(row=2, column=notes_col, value="called - keen, follow up Tuesday")
    ext.save(p)

    store.set_value(2, "Status", "contacted")  # Kairo's own write, after the edit

    reopened = ExcelStore(p, create_if_missing=False)
    _, values, _ = next(reopened.all_rows())
    assert values["Notes"] == "called - keen, follow up Tuesday"
    assert values["Status"] == "contacted"


def test_mailto_prefix_is_read_as_a_bare_address(tmp_path):
    p = _make(tmp_path / "leads.xlsx", ["Name", "Company", "Email", "Phone"],
              ["Ann", "Acme", "mailto:osibgn@gmail.com", "1"])
    store = ExcelStore(p, create_if_missing=False)
    _, values, _ = next(store.all_rows())
    assert values["Email"] == "osibgn@gmail.com"
    assert "osibgn@gmail.com" in store.existing_emails()
    # ... and it's a real send candidate now, not skipped as junk
    assert any(v["Email"] == "osibgn@gmail.com" for _, v in store.rows())


def test_rows_skips_an_unfixable_email(tmp_path):
    p = _make(tmp_path / "leads.xlsx", ["Name", "Company", "Email", "Phone"],
              ["Ann", "Acme", "not an address", "1"])
    store = ExcelStore(p, create_if_missing=False)
    assert list(store.rows()) == []          # never emailed
    assert len(list(store.all_rows())) == 1  # still visible on the Leads page


def test_header_with_trailing_space_still_maps(tmp_path):
    p = _make(tmp_path / "leads.xlsx", ["Name", "Company", "Email ", "Notes "],
              ["Jane", "Acme", "jane@acme.test", "hi"])
    store = ExcelStore(p, create_if_missing=False)
    _, values, _ = next(store.all_rows())
    assert values["Email"] == "jane@acme.test"
    assert values["Notes"] == "hi"


def test_initialize_persists_headers(tmp_path):
    p = tmp_path / "blank.xlsx"
    openpyxl.Workbook().save(p)
    ExcelStore(p, create_if_missing=False).initialize()
    headers = [c.value for c in openpyxl.load_workbook(p).active[1] if c.value]
    assert "Email" in headers and "Name" in headers


def test_add_lead_writes_website_address_only_when_mapped(tmp_path):
    p = _make(tmp_path / "leads.xlsx", ["Name", "Company", "Email", "Phone"],
              ["Jane", "Acme", "jane@acme.test", "1"])
    store = ExcelStore(p)
    store.add_lead({"Company": "NoWeb", "Email": "a@noweb.test",
                    "Website": "https://noweb.test", "Address": "Somewhere"})
    headers = [c.value for c in openpyxl.load_workbook(p).active[1]]
    assert "Website" not in headers and "Address" not in headers

    p2 = _make(tmp_path / "leads2.xlsx", ["Company", "Email", "Website", "Address"],
               ["Acme", "jane@acme.test", "", ""])
    store2 = ExcelStore(p2)
    idx = store2.add_lead({"Company": "WithWeb", "Email": "b@withweb.test",
                           "Website": "https://withweb.test", "Address": "Main St"})
    assert idx is not None
    ws = openpyxl.load_workbook(p2).active
    written = {c.value for c in ws[idx]}
    assert "https://withweb.test" in written and "Main St" in written


def test_remove_rows_deletes_and_keeps_header_and_others(tmp_path):
    p = _make(tmp_path / "leads.xlsx", ["Name", "Company", "Email", "Phone"],
              ["A", "ACo", "a@a.test", "1"],
              ["B", "BCo", "b@b.test", "2"],
              ["C", "CCo", "c@c.test", "3"])
    store = ExcelStore(p)
    i2 = store.add_lead({"Name": "D", "Company": "DCo", "Email": "d@d.test"})
    i1 = store.add_lead({"Name": "E", "Company": "ECo", "Email": "e@e.test"})
    assert (i2, i1) == (5, 6)

    # out-of-range and the header row are ignored
    n = store.remove_rows([i1, i2, 1, 999])
    assert n == 2

    reopened = ExcelStore(p)
    emails = {v["Email"] for _, v, _ in reopened.all_rows()}
    assert emails == {"a@a.test", "b@b.test", "c@c.test"}
    on_disk = [c.value for c in openpyxl.load_workbook(p).active[1]]
    assert on_disk[:4] == ["Name", "Company", "Email", "Phone"]


def test_remove_rows_noop_on_empty(tmp_path):
    p = _make(tmp_path / "leads.xlsx", ["Name", "Email"], ["A", "a@a.test"])
    store = ExcelStore(p)
    assert store.remove_rows([]) == 0
    assert store.remove_rows([1]) == 0


def test_locked_file_reads_rows_via_temp_copy_fallback(tmp_path, monkeypatch):
    p = _make(tmp_path / "leads.xlsx", ["Name", "Company", "Email", "Phone"],
              ["Jane", "Acme", "jane@acme.test", "1"])

    real_load = openpyxl.load_workbook
    state = {"first": True}

    def flaky_load(*a, **k):
        if state["first"]:
            state["first"] = False
            raise PermissionError("open in Excel")
        return real_load(*a, **k)

    monkeypatch.setattr(openpyxl, "load_workbook", flaky_load)
    store = ExcelStore(p, create_if_missing=False)
    assert store.read_only_fallback is True
    _, v = list(store.rows())[0]
    assert v["Email"] == "jane@acme.test"


def test_read_only_fallback_defaults_false(tmp_path):
    p = _make(tmp_path / "leads.xlsx", ["Name", "Email"], ["Jane", "jane@acme.test"])
    assert ExcelStore(p).read_only_fallback is False


def test_locked_uncopyable_primary_falls_back_to_bak(tmp_path, monkeypatch):
    p = _make(tmp_path / "leads.xlsx", ["Name", "Company", "Email", "Phone"],
              ["Jane", "Acme", "jane@acme.test", "1"])
    # First open makes the .bak snapshot.
    ExcelStore(p, create_if_missing=False)

    real_load = openpyxl.load_workbook

    def load(src, *a, **k):
        if str(src).endswith("leads.xlsx"):
            raise zipfile.BadZipFile("truly locked")
        return real_load(src, *a, **k)

    real_copy = shutil.copy2

    def copy(src, dst, *a, **k):
        if str(src).endswith("leads.xlsx"):
            raise OSError("bytes are locked")
        return real_copy(src, dst, *a, **k)

    monkeypatch.setattr(openpyxl, "load_workbook", load)
    monkeypatch.setattr(shutil, "copy2", copy)

    store = ExcelStore(p, create_if_missing=False)
    assert store.read_only_fallback is True
    _, v = list(store.rows())[0]
    assert v["Email"] == "jane@acme.test"


def test_locked_primary_with_no_bak_raises_locked(tmp_path, monkeypatch):
    p = _make(tmp_path / "leads.xlsx", ["Name", "Email"], ["Jane", "jane@acme.test"])

    real_load = openpyxl.load_workbook

    def load(src, *a, **k):
        if str(src).endswith("leads.xlsx"):
            raise PermissionError("open in Excel")
        return real_load(src, *a, **k)

    real_copy = shutil.copy2

    def copy(src, dst, *a, **k):
        if str(src).endswith("leads.xlsx"):
            raise OSError("bytes are locked")
        return real_copy(src, dst, *a, **k)

    monkeypatch.setattr(openpyxl, "load_workbook", load)
    monkeypatch.setattr(shutil, "copy2", copy)

    with pytest.raises(ExcelFileLocked):
        ExcelStore(p, create_if_missing=False)


def test_write_still_raises_locked_even_after_fallback(tmp_path, monkeypatch):
    p = _make(tmp_path / "leads.xlsx", ["Name", "Email"], ["Jane", "jane@acme.test"])

    real_load = openpyxl.load_workbook
    state = {"first": True}

    def flaky_load(*a, **k):
        if state["first"]:
            state["first"] = False
            raise PermissionError("open in Excel")
        return real_load(*a, **k)

    monkeypatch.setattr(openpyxl, "load_workbook", flaky_load)
    store = ExcelStore(p, create_if_missing=False)

    def boom(*a, **k):
        raise PermissionError("still open in Excel")

    monkeypatch.setattr(store.wb, "save", boom)
    with pytest.raises(ExcelFileLocked):
        store.set_value(2, "Notes", "x")


def test_sheet_headers_empty_sheet_returns_empty_without_warning(tmp_path, caplog):
    p = tmp_path / "blank.xlsx"
    openpyxl.Workbook().save(p)
    with caplog.at_level("DEBUG"):
        assert sheet_headers(p) == []
    assert not [r for r in caplog.records if r.levelno >= 30]


def test_sheet_headers_normal_workbook_returns_headers(tmp_path):
    p = _make(tmp_path / "leads.xlsx", ["Name", "Email", "Phone"],
              ["Jane", "jane@acme.test", "1"])
    assert sheet_headers(p) == ["Name", "Email", "Phone"]


def test_data_and_state_column_split_is_exhaustive():
    from kairo.excel_store import LOGICAL_COLUMNS
    assert set(DATA_COLUMNS) | set(STATE_COLUMNS) == set(LOGICAL_COLUMNS)
    assert not (set(DATA_COLUMNS) & set(STATE_COLUMNS))
