from datetime import datetime
from pathlib import Path

import openpyxl

from outreach.blocklist import is_blocked
from outreach.logging_setup import get_logger

log = get_logger("excel_store")

LOGICAL_COLUMNS = [
    "Name",
    "Company",
    "Email",
    "Phone",
    "MeetingDateTime",
    "Status",
    "ColdEmailSentAt",
    "ReminderSentAt",
    "Suppressed",
    "Notes",
    # Reply-handling feature - appended so existing sheets keep their column order.
    "ReplyStatus",     # "", awaiting, yes, no, maybe, question, scheduling, booked
    "LastReplyAt",     # YYYY-MM-DD HH:MM:SS of the most recent inbound reply
    "MeetingEventId",  # Google Calendar event id, set when a booking is approved
]


class ExcelFileLocked(RuntimeError):
    """Raised when clients.xlsx can't be saved, typically because it's open in Excel."""


class ExcelStore:
    def __init__(self, path, column_aliases=None, disallowed_emails=None, disallowed_domains=None):
        self.path = Path(path)
        self.column_aliases = column_aliases or {}
        self.disallowed_emails = disallowed_emails or []
        self.disallowed_domains = disallowed_domains or []
        if not self.path.exists():
            self._create_blank_workbook()
        self.wb = openpyxl.load_workbook(self.path)
        self.ws = self.wb.active
        self._ensure_headers()
        self._index_columns()

    def _header_name(self, logical):
        return self.column_aliases.get(logical, logical)

    def _save(self):
        try:
            self.wb.save(self.path)
        except PermissionError as e:
            raise ExcelFileLocked(
                f"Could not save {self.path} - is it open in Excel? Close it and try again."
            ) from e

    def _create_blank_workbook(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        for col, logical in enumerate(LOGICAL_COLUMNS, start=1):
            ws.cell(row=1, column=col, value=self._header_name(logical))
        try:
            wb.save(self.path)
        except PermissionError as e:
            raise ExcelFileLocked(
                f"Could not create {self.path} - is it open in Excel? Close it and try again."
            ) from e

    def _ensure_headers(self):
        existing = [c.value for c in self.ws[1]]
        changed = False
        for logical in LOGICAL_COLUMNS:
            header = self._header_name(logical)
            if header not in existing:
                self.ws.cell(row=1, column=self.ws.max_column + 1, value=header)
                existing.append(header)
                changed = True
        if changed:
            self._save()

    def _index_columns(self):
        headers = [c.value for c in self.ws[1]]
        self.col_index = {}
        for logical in LOGICAL_COLUMNS:
            header = self._header_name(logical)
            self.col_index[logical] = headers.index(header) + 1  # openpyxl is 1-based

    def _read_row(self, row_idx):
        return {
            logical: self.ws.cell(row=row_idx, column=col).value
            for logical, col in self.col_index.items()
        }

    def get_row(self, row_idx):
        return self._read_row(row_idx)

    def _exclusion_reason(self, values):
        if str(values.get("Suppressed") or "").strip().lower() in ("1", "true", "yes", "y"):
            return "suppressed"
        if values.get("Email") and is_blocked(
            str(values["Email"]), self.disallowed_emails, self.disallowed_domains
        ):
            return "blocked"
        return None

    def rows(self):
        """Yields (row_idx, values) for every row with an email address that isn't suppressed/blocked."""
        for row_idx in range(2, self.ws.max_row + 1):
            values = self._read_row(row_idx)
            if not values.get("Email"):
                continue
            reason = self._exclusion_reason(values)
            if reason:
                if reason == "blocked":
                    log.info(f"Row {row_idx}: {values['Email']} skipped - on the permanent blocklist")
                continue
            yield row_idx, values

    def all_rows(self):
        """Yields (row_idx, values, exclusion_reason) for every row with an email address.

        exclusion_reason is None, "suppressed", or "blocked" - for display purposes
        (e.g. the dashboard's CRM table), unlike rows() which only yields active candidates.
        """
        for row_idx in range(2, self.ws.max_row + 1):
            values = self._read_row(row_idx)
            if not values.get("Email"):
                continue
            yield row_idx, values, self._exclusion_reason(values)

    def set_value(self, row_idx, logical, value):
        col = self.col_index[logical]
        self.ws.cell(row=row_idx, column=col, value=value)
        self._save()  # save immediately so a crash mid-batch never loses state

    def mark_sent(self, row_idx, logical_timestamp_col, when=None):
        when = when or datetime.now()
        self.set_value(row_idx, logical_timestamp_col, when.strftime("%Y-%m-%d %H:%M:%S"))
