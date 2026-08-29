import os
import shutil
import time
from datetime import datetime
from pathlib import Path

import openpyxl

from kairo import column_map
from kairo.blocklist import is_blocked
from kairo.logging_setup import get_logger

log = get_logger("excel_store")

# The lead's own data - supplied by the client, in whatever shape their file has.
# We never add or rename these; column_map.detect() finds them by header text.
# "Priority" is optional: a number the client can put in their sheet to push a
# lead to the front of the send queue when the daily cap bites (kairo/scoring.py).
DATA_COLUMNS = ["Name", "Company", "Email", "Phone", "Priority"]

# Columns this app owns to track outreach state. These are appended to the
# client's file if missing, and are the only columns we ever write to.
STATE_COLUMNS = [
    "MeetingDateTime",
    "Status",
    "ColdEmailSentAt",
    "ReminderSentAt",
    "Suppressed",
    "Notes",
    "ReplyStatus",     # "", awaiting, yes, no, maybe, question, scheduling, booked
    "LastReplyAt",     # YYYY-MM-DD HH:MM:SS of the most recent inbound reply
    "MeetingEventId",  # Google Calendar event id, set when a booking is approved
    "FollowupStage",   # int: how many follow-up nudges have been sent (blank/0 = none)
    "FollowupSentAt",  # YYYY-MM-DD HH:MM:SS of the most recent follow-up
    "ColdMessageId",   # RFC Message-ID of the cold intro, so follow-ups/reminders thread
    "ColdThreadId",    # Gmail threadId of the cold intro
]

LOGICAL_COLUMNS = DATA_COLUMNS + STATE_COLUMNS


class ExcelFileLocked(RuntimeError):
    """Raised when clients.xlsx can't be saved, typically because it's open in Excel."""


class ExcelFileMissing(RuntimeError):
    """Raised when the configured leads file isn't on disk and the caller asked
    us not to silently recreate it (create_if_missing=False). The running app
    passes this so a file the user deleted surfaces as an error on the pages and
    in Diagnostics instead of a fresh empty workbook reappearing."""


def sheet_headers(path):
    """The first-row header text of a workbook, for the Settings mapping UI.
    Returns [] if the file doesn't exist or can't be read."""
    p = Path(path)
    if not p.exists():
        return []
    try:
        wb = openpyxl.load_workbook(p, read_only=True)
        headers = [c.value for c in wb.active[1] if c.value is not None and str(c.value).strip()]
        wb.close()
        return headers
    except Exception as e:  # noqa: BLE001
        log.warning(f"Couldn't read headers from {p}: {e}")
        return []


class ExcelStore:
    def __init__(self, path, column_map=None, disallowed_emails=None,
                 disallowed_domains=None, column_aliases=None, create_if_missing=True):
        self.path = Path(path)
        # column_aliases is the old (wizard-era) name for the same thing.
        self.explicit_map = dict(column_map or column_aliases or {})
        self.disallowed_emails = disallowed_emails or []
        self.disallowed_domains = disallowed_domains or []
        self.email_column_missing = False
        self._headers_dirty = False
        if not self.path.exists():
            if not create_if_missing:
                raise ExcelFileMissing(
                    f"The leads file {self.path} is not on this PC. Kairo will not "
                    f"recreate it - restore the file or pick another one under Settings."
                )
            self._create_blank_workbook()
        else:
            self._refresh_backup()
        self.wb = self._load_workbook()
        self.ws = self.wb.active
        self._select_sheet_with_email()
        self._load_headers()
        self._ensure_headers()
        self._resolve_map()
        self._index_columns()

    # --- setup -----------------------------------------------------------

    @staticmethod
    def _row1_cells(ws):
        """[(column_index, value)] for every non-empty header cell in row 1.

        Keeps the real 1-based column index of each header - a blank column in
        the middle of the sheet must not shift the ones after it, or every read
        past the gap lands in the wrong cell (e.g. a Note typed in Excel never
        showing up on the Leads page)."""
        out = []
        for cell in ws[1]:
            if cell.value is not None and str(cell.value).strip():
                out.append((cell.column, cell.value))
        return out

    @classmethod
    def _row1_headers(cls, ws):
        return [v for _, v in cls._row1_cells(ws)]

    def _load_headers(self):
        """(Re)build self.headers (names, in column order) and self._header_col
        (name -> real 1-based column) from the active worksheet."""
        cells = self._row1_cells(self.ws)
        self.headers = [v for _, v in cells]
        self._header_col = {}
        for col, value in cells:
            self._header_col.setdefault(value, col)

    def _load_workbook(self):
        """Open the workbook, retrying once on a transient read error. A genuine
        lock (file open exclusively) still surfaces as ExcelFileLocked."""
        last_err = None
        for attempt in range(2):
            try:
                return openpyxl.load_workbook(self.path)
            except PermissionError as e:
                raise ExcelFileLocked(
                    f"Could not open {self.path} - is it open in Excel? Close it and try again."
                ) from e
            except Exception as e:  # noqa: BLE001
                last_err = e
                if attempt == 0:
                    time.sleep(0.1)
        raise RuntimeError(f"Could not read {self.path}: {last_err}")

    def _refresh_backup(self):
        """Best-effort snapshot so a crash or corruption mid-save leaves the
        client a recoverable .bak. web/app.py builds a store per request, so only
        refresh the copy when it's missing or over 24h stale."""
        bak = self.path.with_suffix(self.path.suffix + ".bak")
        for attempt in range(2):
            try:
                if not bak.exists() or time.time() - os.path.getmtime(bak) > 86400:
                    shutil.copy2(self.path, bak)
                return
            except PermissionError:
                return  # locked for reading right now; a stale .bak is fine
            except OSError:
                if attempt == 0:
                    time.sleep(0.1)

    def _select_sheet_with_email(self):
        """The active sheet is usually the leads. If it has no detectable Email
        column but another worksheet does, switch to that one. Only picks the
        sheet - _load_headers() then reads headers off whichever sheet wins."""
        def has_email(ws):
            names = self._row1_headers(ws)
            return "Email" in column_map.detect([h for h in names if h not in STATE_COLUMNS])

        if has_email(self.ws):
            return
        for ws in self.wb.worksheets:
            if ws is not self.ws and has_email(ws):
                log.info(
                    f"Using sheet {ws.title!r} - it has an Email column and the active sheet doesn't"
                )
                self.ws = ws
                return

    def _save(self):
        try:
            tmp = self.path.with_name(self.path.name + ".tmp")
            self.wb.save(tmp)
            os.replace(tmp, self.path)
        except PermissionError as e:
            raise ExcelFileLocked(
                f"Could not save {self.path} - is it open in Excel? Close it and try again."
            ) from e

    def _create_blank_workbook(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        for col, logical in enumerate(LOGICAL_COLUMNS, start=1):
            ws.cell(row=1, column=col, value=logical)
        try:
            wb.save(self.path)
        except PermissionError as e:
            raise ExcelFileLocked(
                f"Could not create {self.path} - is it open in Excel? Close it and try again."
            ) from e

    def _ensure_headers(self):
        """Register any missing STATE columns. Never touches the client's data
        columns, and never writes during construction - the header cells are held
        in memory (openpyxl keeps them until .save()) and only land on disk when a
        real write (set_value / add_lead) persists the workbook."""
        for logical in STATE_COLUMNS:
            if logical not in self.headers:
                col = self.ws.max_column + 1
                self.ws.cell(row=1, column=col, value=logical)
                self.headers.append(logical)
                self._header_col[logical] = col
                self._headers_dirty = True

    def _resolve_map(self):
        """Merge auto-detection with the explicit config map (config wins).

        Detection runs only over the client's own headers - the STATE columns we
        appended must not be candidates (e.g. 'ColdEmailSentAt' contains 'email')."""
        data_headers = [h for h in self.headers if h not in STATE_COLUMNS]
        detected = column_map.detect(data_headers)
        self.effective_map = {**detected, **self.explicit_map}

    # --- column indexing -----------------------------------------------

    def _col_for_header(self, header):
        return self._header_col.get(header)

    def _index_columns(self):
        self.col_index = {}

        for logical in DATA_COLUMNS:
            target = self.effective_map.get(logical)
            if target is None:
                continue
            if isinstance(target, (list, tuple)):
                cols = [c for c in (self._col_for_header(h) for h in target) if c]
                if len(cols) == 1:
                    self.col_index[logical] = cols[0]
                elif cols:
                    self.col_index[logical] = cols
            else:
                col = self._col_for_header(target)
                if col:
                    self.col_index[logical] = col

        for logical in STATE_COLUMNS:
            col = self._col_for_header(logical)
            if col:
                self.col_index[logical] = col

        self.email_column_missing = "Email" not in self.col_index

    # --- reading ------------------------------------------------------

    def _cell(self, row_idx, col):
        if isinstance(col, list):
            parts = [self.ws.cell(row=row_idx, column=c).value for c in col]
            parts = [str(p).strip() for p in parts if p is not None and str(p).strip()]
            return " ".join(parts) or None
        return self.ws.cell(row=row_idx, column=col).value

    def _read_row(self, row_idx):
        return {logical: self._cell(row_idx, col) for logical, col in self.col_index.items()}

    def get_row(self, row_idx):
        return self._read_row(row_idx)

    def _exclusion_reason(self, values):
        if str(values.get("Suppressed") or "").strip().lower() in ("1", "true", "yes", "y"):
            return "suppressed"
        if values.get("Email") and is_blocked(
            str(values["Email"]), self.disallowed_emails, self.disallowed_domains
        ):
            return "blocked"
        return None

    def rows(self):
        """Yields (row_idx, values) for every row with an email address that isn't suppressed/blocked."""
        if self.email_column_missing:
            return
        for row_idx in range(2, self.ws.max_row + 1):
            values = self._read_row(row_idx)
            if not values.get("Email"):
                continue
            reason = self._exclusion_reason(values)
            if reason:
                if reason == "blocked":
                    log.info(f"Row {row_idx}: {values['Email']} skipped - on the permanent blocklist")
                continue
            yield row_idx, values

    def all_rows(self):
        """Yields (row_idx, values, exclusion_reason) for every row with an email address.

        exclusion_reason is None, "suppressed", or "blocked" - for display purposes
        (e.g. the dashboard's CRM table), unlike rows() which only yields active candidates.
        """
        if self.email_column_missing:
            return
        for row_idx in range(2, self.ws.max_row + 1):
            values = self._read_row(row_idx)
            if not values.get("Email"):
                continue
            yield row_idx, values, self._exclusion_reason(values)

    # --- writing (STATE columns only) --------------------------------

    def set_value(self, row_idx, logical, value):
        if logical not in STATE_COLUMNS:
            raise ValueError(f"Refusing to write to non-state column {logical!r}")
        col = self.col_index.get(logical)
        if not isinstance(col, int):
            raise ValueError(f"State column {logical!r} is not present in the sheet")
        self.ws.cell(row=row_idx, column=col, value=value)
        self._save()  # save immediately so a crash mid-batch never loses state

    def mark_sent(self, row_idx, logical_timestamp_col, when=None):
        when = when or datetime.now()
        self.set_value(row_idx, logical_timestamp_col, when.strftime("%Y-%m-%d %H:%M:%S"))

    # --- appending new leads (lead sourcing, BETA) -------------------
    # This is the only non-state write. It's exempt from the state-column guard
    # because it adds a brand-new row - it never mutates a cell the client owns.

    def existing_emails(self):
        """Lower-cased set of every email already in the sheet."""
        return {
            str(v["Email"]).strip().lower()
            for _, v, _ in self.all_rows()
            if str(v.get("Email") or "").strip()
        }

    def _append_col_for(self, logical):
        """Column index to write a new-lead field to. Name/Company/Email/Phone use
        the resolved map; Website/Address only if the sheet already has such a
        column (mapped or a header of that literal name). Never creates a header."""
        col = self.col_index.get(logical)
        if isinstance(col, int):
            return col
        target = self.effective_map.get(logical) or (logical if logical in self.headers else None)
        if isinstance(target, str):
            return self._col_for_header(target)
        return None

    def add_lead(self, data):
        """Append a lead row. Returns the new row index, or None if the email is
        blank / invalid / already present. Only writes fields that are mapped to a
        real column - never invents a header."""
        from kairo.lead_fields import valid_email

        email = str(data.get("Email") or "").strip()
        if not valid_email(email) or email.lower() in self.existing_emails():
            return None

        row_idx = self.ws.max_row + 1
        wrote_any = False
        for logical in ("Name", "Company", "Email", "Phone", "Website", "Address"):
            value = data.get(logical)
            if value in (None, ""):
                continue
            col = self._append_col_for(logical)
            if isinstance(col, int):
                self.ws.cell(row=row_idx, column=col, value=value)
                wrote_any = True
        if not wrote_any:
            return None
        self._save()
        return row_idx

    def remove_rows(self, row_idxs):
        """Delete the given 1-based worksheet rows. Like add_lead this is a
        structural write - it removes whole rows rather than mutating a cell the
        client owns - so it's exempt from the state-column guard. Rows are deleted
        in descending order so earlier indices stay valid. The header row (1) and
        any index outside the sheet are ignored. Returns the number of rows
        deleted."""
        max_row = self.ws.max_row
        targets = sorted(
            {i for i in row_idxs if isinstance(i, int) and 2 <= i <= max_row},
            reverse=True,
        )
        for i in targets:
            self.ws.delete_rows(i, 1)
        if targets:
            self._save()
        return len(targets)
